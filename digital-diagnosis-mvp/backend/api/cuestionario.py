"""
API del Cuestionario de Diagnóstico
Herramienta de Diagnóstico Digital para Unidades Productivas
A1ia Tech - MVP
"""

from fastapi import APIRouter, HTTPException, Depends, status
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from datetime import datetime
import uuid

from ..database.database import get_db
from ..database.models import Pregunta, Diagnostico, Respuesta, Usuario, Dimension, Subdimension
from ..config import settings
from .auth import get_current_user

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import json

router = APIRouter()

class PreguntaResponse(BaseModel):
    id: str
    dimension_id: int
    texto_pregunta: str
    tipo_pregunta: str
    opciones_respuesta: List[Dict]

@router.get("/estructura")
async def obtener_estructura_cuestionario(db: Session = Depends(get_db)):
    """Obtener estructura del cuestionario"""
    
    preguntas = db.query(Pregunta).filter(Pregunta.activa == True).limit(10).all()
    
    preguntas_formateadas = []
    for pregunta in preguntas:
        preguntas_formateadas.append(PreguntaResponse(
            id=pregunta.id,
            dimension_id=pregunta.dimension_id,
            texto_pregunta=pregunta.texto_pregunta,
            tipo_pregunta=pregunta.tipo_pregunta,
            opciones_respuesta=pregunta.opciones_respuesta
        ))
    
    return {
        "total_preguntas": len(preguntas_formateadas),
        "preguntas": preguntas_formateadas
    }

@router.post("/iniciar")
async def iniciar_diagnostico(
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Iniciar nuevo diagnóstico"""
    
    nuevo_diagnostico = Diagnostico(
        id=str(uuid.uuid4()),
        usuario_id=current_user.id,
        fecha_inicio=datetime.utcnow(),
        estado="EN_PROGRESO"
    )
    
    db.add(nuevo_diagnostico)
    db.commit()
    db.refresh(nuevo_diagnostico)
    
    return {
        "diagnostico_id": nuevo_diagnostico.id,
        "estado": nuevo_diagnostico.estado,
        "fecha_inicio": nuevo_diagnostico.fecha_inicio.isoformat()
    }

@router.get("/plantilla-excel")
async def descargar_plantilla_excel(db: Session = Depends(get_db)):
    """
    Genera y descarga una plantilla Excel para que el asesor llene las respuestas del cuestionario
    """
    try:
        print("=== INICIO GENERACIÓN PLANTILLA ===")
        
        # Obtener todas las preguntas de la base de datos
        preguntas = db.query(Pregunta).order_by(Pregunta.id).all()
        print(f"Preguntas obtenidas: {len(preguntas)}")
        
        # Debug: Veamos cómo están las opciones en la primera pregunta
        if preguntas:
            primera_pregunta = preguntas[0]
            print(f"Tipo de opciones_respuesta: {type(primera_pregunta.opciones_respuesta)}")
            print(f"Contenido opciones_respuesta: {primera_pregunta.opciones_respuesta}")
            print(f"Longitud: {len(primera_pregunta.opciones_respuesta) if primera_pregunta.opciones_respuesta else 'None'}")
        
        if not preguntas:
            raise HTTPException(status_code=404, detail="No se encontraron preguntas en la base de datos")
        
        # Crear un nuevo libro de Excel
        print("Creando libro Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuestionario Diagnóstico"
        
        # Configurar encabezados
        headers = [
            "ID Pregunta",
            "Dimensión", 
            "Pregunta",
            "Opciones Disponibles",
            "Respuesta Seleccionada"
        ]
        
        print("Escribiendo encabezados...")
        # Escribir encabezados con formato
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="151F3B", end_color="151F3B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        print("Procesando preguntas...")
        # Llenar las filas con las preguntas
        for idx, pregunta in enumerate(preguntas, 1):
            fila = idx + 1  # Empezar en fila 2 (después de encabezados)
            
            print(f"Procesando pregunta {idx}: {pregunta.id}")
            
            # ID de pregunta
            ws.cell(row=fila, column=1, value=pregunta.id)
            
            # Dimensión
            print(f"Obteniendo dimensión para: {pregunta.dimension_id}")
            dimension_nombre = obtener_nombre_dimension(pregunta.dimension_id)
            ws.cell(row=fila, column=2, value=dimension_nombre)
            
            # Pregunta
            ws.cell(row=fila, column=3, value=pregunta.texto_pregunta)
            
            # Opciones disponibles
            print(f"Procesando opciones de respuesta...")
            try:
                # Manejar diferentes formatos posibles
                opciones_raw = pregunta.opciones_respuesta
                
                if isinstance(opciones_raw, str):
                    # Si es string, intentar parsear JSON
                    opciones = json.loads(opciones_raw)
                elif isinstance(opciones_raw, list):
                    # Si ya es lista, usar directamente
                    opciones = opciones_raw
                else:
                    # Si es otro tipo, convertir a string y parsear
                    opciones = json.loads(str(opciones_raw))
                
                # Crear texto descriptivo de opciones
                if isinstance(opciones, list) and len(opciones) > 0:
                    opciones_texto = "\n".join([f"{op.get('texto', 'Sin texto')} ({op.get('valor', 0)} pts)" for op in opciones])
                else:
                    opciones_texto = "Sin opciones definidas"
                    
                cell_opciones = ws.cell(row=fila, column=4, value=opciones_texto)
                cell_opciones.alignment = Alignment(wrap_text=True, vertical="top")
                
            except Exception as e:
                print(f"Error procesando opciones para pregunta {pregunta.id}: {e}")
                print(f"Tipo de dato: {type(pregunta.opciones_respuesta)}")
                print(f"Contenido: {pregunta.opciones_respuesta}")
                ws.cell(row=fila, column=4, value=f"Error: {str(e)[:50]}")
            
            # Columna vacía para respuesta (el asesor la llenará)
            ws.cell(row=fila, column=5, value="")
        
        # Ajustar altura de filas para mostrar texto multilínea
        print("Ajustando altura de filas...")
        for row in range(2, len(preguntas) + 2):
            ws.row_dimensions[row].height = 60
        
        print("Ajustando columnas...")
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 40
        
        print("Guardando en memoria...")
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print("Creando respuesta...")
        # Crear respuesta de descarga
        response = StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = "attachment; filename=plantilla_cuestionario_diagnostico.xlsx"
        
        print("=== PLANTILLA GENERADA EXITOSAMENTE ===")
        return response
        
    except Exception as e:
        print(f"=== ERROR EN GENERACIÓN PLANTILLA: {e} ===")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generando plantilla: {str(e)}")

def obtener_nombre_dimension(dimension_id):
    """Helper para obtener nombres legibles de las dimensiones"""
    dimensiones = {
        1: "Estrategia y Liderazgo Digital",
        2: "Procesos, Operaciones y Datos", 
        3: "Marketing y Experiencia del Cliente",
        4: "Tecnología e Infraestructura",
        5: "Cultura y Talento Humano",
        6: "Ciberseguridad y Gobernanza"
    }
    return dimensiones.get(int(dimension_id), f"Dimensión {dimension_id}")