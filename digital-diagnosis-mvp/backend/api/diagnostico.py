"""
API de Diagnóstico básica
Herramienta de Diagnóstico Digital para Unidades Productivas
A1ia Tech - MVP
"""

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from ..database.database import get_db
from ..ai_engine.gemini_client import GeminiClient
from .auth import get_current_user

from fastapi import UploadFile, File, HTTPException, Form
from pydantic import BaseModel
from datetime import datetime
from typing import Dict, Any
import json
import uuid
import io
from openpyxl import load_workbook
from ..database.models import Pregunta, Diagnostico, Respuesta, Usuario

router = APIRouter()

@router.get("/test-ia")
async def test_conexion_ia():
    """Probar conexión con Gemini"""
    try:
        gemini_client = GeminiClient()
        conexion_ok = gemini_client.test_connection()
        
        return {
            "gemini_disponible": conexion_ok,
            "modelo": "gemini-pro",
            "estado": "OK" if conexion_ok else "ERROR"
        }
    except Exception as e:
        return {
            "gemini_disponible": False,
            "error": str(e),
            "estado": "ERROR"
        }

@router.post("/generar-dofa-test")
async def generar_dofa_test():
    """Test de generación DOFA"""
    try:
        gemini_client = GeminiClient()
        
        puntajes_test = {
            "global": 45.5,
            "dimensiones": {1: 30, 2: 60, 3: 40, 4: 50, 5: 35, 6: 55}
        }
        
        dofa = gemini_client.generar_analisis_dofa({}, puntajes_test)
        
        return {
            "status": "success",
            "dofa_generado": dofa
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e)
        }

@router.post("/procesar-excel")
async def procesar_archivo_excel(
    archivo: UploadFile = File(...),
    up_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Procesa el archivo Excel llenado por el asesor y genera el diagnóstico automático con IA
    """
    try:
        print("=== INICIO PROCESAMIENTO ARCHIVO EXCEL ===")
        
        # Validar tipo de archivo
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
        
        # Leer archivo Excel
        print("Leyendo archivo Excel...")
        contents = await archivo.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Extraer respuestas del Excel
        respuestas = {}
        errores_validacion = []
        
        print("Extrayendo respuestas...")
        # Empezar desde fila 2 (después de encabezados)
        for row in range(2, ws.max_row + 1):
            id_pregunta = ws.cell(row=row, column=1).value
            respuesta_seleccionada = ws.cell(row=row, column=5).value
            
            if id_pregunta and respuesta_seleccionada:
                # Validar que la pregunta existe en la base de datos
                pregunta = db.query(Pregunta).filter(Pregunta.id == id_pregunta).first()
                if not pregunta:
                    errores_validacion.append(f"Pregunta {id_pregunta} no encontrada")
                    continue
                
                # Validar que la respuesta es válida para esa pregunta
                try:
                    opciones = json.loads(pregunta.opciones_respuesta) if isinstance(pregunta.opciones_respuesta, str) else pregunta.opciones_respuesta
                    opciones_validas = [op['texto'] for op in opciones]
                    
                    # Limpiar respuesta (quitar comillas extra si las hay)
                    respuesta_limpia = str(respuesta_seleccionada).strip().strip('"').strip("'")

                    if respuesta_limpia not in opciones_validas:
                        errores_validacion.append(f"Respuesta '{respuesta_limpia}' no válida para pregunta {id_pregunta}. Opciones: {opciones_validas}")
                        continue

                    # Obtener el valor numérico de la respuesta
                    valor_respuesta = next((op['valor'] for op in opciones if op['texto'] == respuesta_limpia), 0)

                    # Asegurar que el valor no sea None
                    if valor_respuesta is None:
                        valor_respuesta = 0
                    
                    respuestas[id_pregunta] = {
                        'texto': respuesta_limpia,
                        'valor': valor_respuesta,
                        'dimension_id': pregunta.dimension_id
                    }
                    
                except Exception as e:
                    errores_validacion.append(f"Error procesando pregunta {id_pregunta}: {str(e)}")
        
        print(f"Respuestas válidas procesadas: {len(respuestas)}")
        
        if errores_validacion:
            print(f"Errores de validación: {errores_validacion}")
            raise HTTPException(
                status_code=400, 
                detail=f"Errores en el archivo: {'; '.join(errores_validacion[:5])}"
            )
        
        if len(respuestas) == 0:
            raise HTTPException(status_code=400, detail="No se encontraron respuestas válidas en el archivo")
        
        # Calcular puntajes por dimensión
        print("Calculando puntajes...")
        puntajes = calcular_puntajes_diagnostico(respuestas)
        
        # Crear nuevo diagnóstico
        print("Creando diagnóstico...")
        nuevo_diagnostico = Diagnostico(
            id=str(uuid.uuid4()),
            usuario_id=current_user.id,
            fecha_inicio=datetime.utcnow(),
            puntaje_global=puntajes['global'],
            puntajes_dimensiones=json.dumps(puntajes['dimensiones']),
            estado="PROCESANDO_IA"
        )
        
        db.add(nuevo_diagnostico)
        db.flush()
        
        # Guardar respuestas en la base de datos
        print("Guardando respuestas...")
        for id_pregunta, respuesta_data in respuestas.items():
            nueva_respuesta = Respuesta(
                diagnostico_id=nuevo_diagnostico.id,
                pregunta_id=id_pregunta,
                valor_calculado=respuesta_data['valor'],
                respuesta_seleccionada=respuesta_data['texto']
            )
            db.add(nueva_respuesta)
        
        # Generar análisis DOFA con IA real
        print("Generando análisis DOFA con IA...")
        try:
            gemini_client = GeminiClient()
            dofa_analisis = gemini_client.generar_analisis_dofa(respuestas, puntajes)
            
            # Actualizar diagnóstico con DOFA y estado final
            nuevo_diagnostico.analisis_dofa = json.dumps(dofa_analisis)
            nuevo_diagnostico.estado = "COMPLETADO"
            nuevo_diagnostico.fecha_finalizacion = datetime.utcnow()
            
        except Exception as e:
            print(f"Error generando DOFA: {e}")
            nuevo_diagnostico.estado = "COMPLETADO_SIN_IA"
            nuevo_diagnostico.fecha_finalizacion = datetime.utcnow()
        
        db.commit()
        
        print("=== ARCHIVO PROCESADO EXITOSAMENTE ===")
        
        return {
            "success": True,
            "diagnostico_id": nuevo_diagnostico.id,
            "respuestas_procesadas": len(respuestas),
            "puntaje_global": puntajes['global'],
            "puntajes_dimensiones": puntajes['dimensiones'],
            "dofa_generado": nuevo_diagnostico.analisis_dofa is not None,
            "message": f"Diagnóstico completado exitosamente. {len(respuestas)} respuestas procesadas."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"=== ERROR PROCESANDO ARCHIVO: {e} ===")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

def calcular_puntajes_diagnostico(respuestas):
    """
    Calcula los puntajes por dimensión basado en las respuestas
    """
    # Agrupar respuestas por dimensión
    puntajes_por_dimension = {}
    for id_pregunta, respuesta_data in respuestas.items():
        dimension_id = respuesta_data['dimension_id']
        valor = respuesta_data['valor']
        
        # Filtrar valores None o inválidos
        if valor is not None and isinstance(valor, (int, float)):
            if dimension_id not in puntajes_por_dimension:
                puntajes_por_dimension[dimension_id] = []
            puntajes_por_dimension[dimension_id].append(valor)
    
    # Calcular promedio por dimensión
    puntajes_dimensiones = {}
    for dimension_id, valores in puntajes_por_dimension.items():
        if valores:  # Solo si hay valores válidos
            puntajes_dimensiones[dimension_id] = round(sum(valores) / len(valores), 2)
        else:
            puntajes_dimensiones[dimension_id] = 0
    
    # Calcular puntaje global
    if puntajes_dimensiones:
        valores_dimensiones = [v for v in puntajes_dimensiones.values() if v is not None]
        puntaje_global = round(sum(valores_dimensiones) / len(valores_dimensiones), 2) if valores_dimensiones else 0
    else:
        puntaje_global = 0
    
    return {
        'global': puntaje_global,
        'dimensiones': puntajes_dimensiones
    }

class AnalisisDimensionRequest(BaseModel):
    dimension: str
    puntaje: int
    up_nombre: str
    up_sector: str
    diagnostico_id: str = None
    puntajes_todas_dimensiones: Dict[str, int] = {}

@router.post("/generar-analisis-dimension")
async def generar_analisis_dimension(
    request: AnalisisDimensionRequest,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Genera análisis personalizado usando Gemini AI para una dimensión específica
    """
    try:
        print(f"=== GENERANDO ANÁLISIS GEMINI PARA {request.dimension} ===")
        print(f"UP: {request.up_nombre} | Sector: {request.up_sector} | Puntaje: {request.puntaje}")
        
        # Verificar autenticación
        if not current_user:
            raise HTTPException(status_code=401, detail="Usuario no autenticado")
        
        # Validar datos de entrada
        if not request.dimension or not request.up_nombre:
            raise HTTPException(status_code=400, detail="Datos incompletos en la solicitud")
        
        if request.puntaje < 0 or request.puntaje > 100:
            raise HTTPException(status_code=400, detail="Puntaje debe estar entre 0 y 100")
        
        # Inicializar cliente Gemini
        gemini_client = GeminiClient()
        
        # Verificar conexión con Gemini
        if not gemini_client.test_connection():
            raise HTTPException(status_code=503, detail="Servicio de IA no disponible")
        
        print("✅ Cliente Gemini inicializado y conectado")
        
        # Generar análisis con Gemini AI
        print("🤖 Llamando a Gemini AI...")
        
        analisis = gemini_client.generar_analisis_dimension(
            dimension=request.dimension,
            puntaje=request.puntaje,
            up_nombre=request.up_nombre,
            up_sector=request.up_sector,
            puntajes_generales=request.puntajes_todas_dimensiones
        )
        
        print("✅ Análisis generado exitosamente")
        
        return {
            "success": True,
            "analisis": analisis,
            "dimension": request.dimension,
            "puntaje": request.puntaje,
            "up_nombre": request.up_nombre,
            "timestamp": datetime.utcnow().isoformat(),
            "generado_por": "Gemini AI"
        }
        
    except HTTPException:
        # Re-lanzar HTTPExceptions tal como están
        raise
    except Exception as e:
        print(f"❌ Error generando análisis: {e}")
        import traceback
        traceback.print_exc()
        
        # En caso de error, devolver análisis básico
        analisis_fallback = generar_analisis_basico_fallback(
            request.dimension, 
            request.puntaje, 
            request.up_nombre
        )
        
        return {
            "success": False,
            "analisis": analisis_fallback,
            "dimension": request.dimension,
            "puntaje": request.puntaje,
            "error": "Error de IA - usando análisis básico",
            "timestamp": datetime.utcnow().isoformat(),
            "generado_por": "Sistema básico"
        }

def generar_analisis_basico_fallback(dimension: str, puntaje: int, up_nombre: str) -> str:
    """Genera análisis básico en caso de fallo completo del sistema"""
    
    if puntaje >= 80:
        nivel = "excelente"
        descripcion = "muestra un dominio avanzado"
        recomendacion = "mantener las buenas prácticas actuales y buscar oportunidades de liderazgo"
    elif puntaje >= 60:
        nivel = "bueno"
        descripcion = "tiene una base sólida con oportunidades claras"
        recomendacion = "enfocarse en optimizar procesos específicos identificados"
    elif puntaje >= 40:
        nivel = "regular"
        descripcion = "muestra desarrollo parcial que requiere atención"
        recomendacion = "desarrollar capacidades fundamentales con un plan estructurado"
    else:
        nivel = "básico"
        descripcion = "está en etapa inicial y necesita desarrollo prioritario"
        recomendacion = "implementar mejoras estructurales con soporte especializado"
    
    return f"""**SITUACIÓN ACTUAL**
{up_nombre} presenta un nivel {nivel} en {dimension} con {puntaje}/100 puntos. La empresa {descripcion} en esta dimensión crítica para la transformación digital.

**ÁREAS DE MEJORA**
Se identifican oportunidades específicas de crecimiento en {dimension} que pueden generar mayor impacto en el desempeño del negocio.

**RECOMENDACIONES PRIORITARIAS**  
Es recomendable {recomendacion} para avanzar efectivamente en la madurez digital de {dimension}."""